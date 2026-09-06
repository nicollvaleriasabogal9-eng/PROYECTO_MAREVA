from repositories.reporte_repository import ReporteRepository
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import io


class ReporteServices:

    def __init__(self):
        self.repo = ReporteRepository()

    # RF-89: Reporte de reservas por período
    def generar_reporte_reservas_periodo(self, fecha_inicio, fecha_fin, estado=None):
        """Genera reporte de reservas en un período"""
        datos = self.repo.obtener_reservas_por_periodo(fecha_inicio, fecha_fin, estado)
        
        total_reservas = len(datos)
        ingresos_totales = sum(r["precio"] for r in datos)
        personas_totales = sum(r["cant_adultos"] + r["cant_menores"] for r in datos)
        
        estados_conteo = {}
        for r in datos:
            estado = r["estado"]
            estados_conteo[estado] = estados_conteo.get(estado, 0) + 1
        
        return {
            "datos": datos,
            "estadisticas": {
                "total_reservas": total_reservas,
                "ingresos_totales": ingresos_totales,
                "personas_totales": personas_totales,
                "estados_conteo": estados_conteo,
                "promedio_por_reserva": ingresos_totales / total_reservas if total_reservas > 0 else 0
            }
        }

    # RF-90: Reporte de paquetes más reservados
    def generar_reporte_paquetes_top(self, fecha_inicio, fecha_fin):
        """Genera ranking de paquetes más reservados"""
        datos = self.repo.obtener_paquetes_mas_reservados(fecha_inicio, fecha_fin)
        
        total_reservas = sum(p["total_reservas"] for p in datos)
        total_personas = sum(p["total_personas"] for p in datos)
        ingreso_total = sum(p["total_reservas"] * p["precio_promedio"] for p in datos)
        
        return {
            "datos": datos,
            "estadisticas": {
                "total_paquetes": len(datos),
                "total_reservas": total_reservas,
                "total_personas": total_personas,
                "ingreso_total": ingreso_total,
                "paquete_top": datos[0] if datos else None
            }
        }

    # RF-91: Reporte de reservas canceladas
    def generar_reporte_canceladas(self, fecha_inicio, fecha_fin):
        """Genera reporte de reservas canceladas"""
        datos = self.repo.obtener_reservas_canceladas(fecha_inicio, fecha_fin)
        
        total_canceladas = len(datos)
        valor_cancelado = sum(r["precio"] for r in datos)
        personas_afectadas = 0  # Podría sumarse cant_adultos + cant_menores si lo tuviera
        
        return {
            "datos": datos,
            "estadisticas": {
                "total_canceladas": total_canceladas,
                "valor_cancelado": valor_cancelado,
                "promedio_cancelacion": valor_cancelado / total_canceladas if total_canceladas > 0 else 0
            }
        }

    # RF-92: Reporte de ingresos esperados
    def generar_reporte_ingresos(self, fecha_inicio, fecha_fin):
        """Genera reporte de ingresos esperados"""
        datos = self.repo.obtener_ingresos_esperados(fecha_inicio, fecha_fin)
        return datos

    # RF-93: Reporte de destinos por temporada
    def generar_reporte_destinos_temporada(self, fecha_inicio, fecha_fin):
        """Genera reporte de destinos más demandados por temporada"""
        datos = self.repo.obtener_destinos_por_temporada(fecha_inicio, fecha_fin)
        
        # Agrupar por destino
        destinos_agrupados = {}
        for item in datos:
            nombre = item["nombre"]
            if nombre not in destinos_agrupados:
                destinos_agrupados[nombre] = {
                    "id_destino": item["id_destino"],
                    "nombre": nombre,
                    "temporadas": [],
                    "total_reservas_general": 0,
                    "total_personas_general": 0
                }
            destinos_agrupados[nombre]["temporadas"].append(item)
            destinos_agrupados[nombre]["total_reservas_general"] += item["cantidad_reservas"]
            destinos_agrupados[nombre]["total_personas_general"] += item["total_personas"]
        
        # Ordenar por total de reservas
        destinos_ordenados = sorted(
            destinos_agrupados.values(),
            key=lambda x: x["total_reservas_general"],
            reverse=True
        )
        
        return {
            "datos": destinos_ordenados,
            "estadisticas": {
                "total_destinos": len(destinos_ordenados),
                "total_reservas": sum(d["total_reservas_general"] for d in destinos_ordenados),
                "total_personas": sum(d["total_personas_general"] for d in destinos_ordenados)
            }
        }

    # Exportar a Excel
    def exportar_excel_reservas_periodo(self, fecha_inicio, fecha_fin, estado=None):
        """Exporta reporte de reservas a Excel"""
        reporte = self.generar_reporte_reservas_periodo(fecha_inicio, fecha_fin, estado)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Reservas"
        
        # Estilos
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws['A1'] = "REPORTE DE RESERVAS POR PERÍODO"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:L1')
        
        # Información del período
        ws['A2'] = f"Período: {fecha_inicio} a {fecha_fin}"
        ws['A3'] = f"Fecha de generación: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # Estadísticas
        ws['A5'] = "ESTADÍSTICAS GENERALES"
        ws['A5'].font = Font(bold=True, size=11)
        ws['A6'] = f"Total de reservas: {reporte['estadisticas']['total_reservas']}"
        ws['A7'] = f"Ingresos totales: ${reporte['estadisticas']['ingresos_totales']:,.2f}"
        ws['A8'] = f"Total de personas: {reporte['estadisticas']['personas_totales']}"
        ws['A9'] = f"Promedio por reserva: ${reporte['estadisticas']['promedio_por_reserva']:,.2f}"
        
        # Encabezados de tabla
        headers = ["ID", "Código", "Fecha Viaje", "Fecha Reserva", "Estado", "Adultos", 
                  "Menores", "Plan", "Cliente", "Correo", "Paquete", "Precio"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=11, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Datos
        for row, reserva in enumerate(reporte['datos'], 12):
            ws.cell(row=row, column=1).value = reserva['id_reserva']
            ws.cell(row=row, column=2).value = reserva['codigo_unico']
            ws.cell(row=row, column=3).value = reserva['fecha_viaje']
            ws.cell(row=row, column=4).value = reserva['fecha_reserva']
            ws.cell(row=row, column=5).value = reserva['estado']
            ws.cell(row=row, column=6).value = reserva['cant_adultos']
            ws.cell(row=row, column=7).value = reserva['cant_menores']
            ws.cell(row=row, column=8).value = reserva['plan']
            ws.cell(row=row, column=9).value = reserva['cliente_nombre']
            ws.cell(row=row, column=10).value = reserva['cliente_correo']
            ws.cell(row=row, column=11).value = reserva['paquete_nombre']
            ws.cell(row=row, column=12).value = reserva['precio']
            
            # Aplicar bordes y alineación
            for col in range(1, 13):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                if col == 12:  # Columna de precio
                    cell.number_format = '$#,##0.00'
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 8
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 20
        ws.column_dimensions['J'].width = 25
        ws.column_dimensions['K'].width = 20
        ws.column_dimensions['L'].width = 12
        
        # Guardar en BytesIO
        excel_io = io.BytesIO()
        wb.save(excel_io)
        excel_io.seek(0)
        
        return excel_io

    def exportar_excel_paquetes_top(self, fecha_inicio, fecha_fin):
        """Exporta ranking de paquetes a Excel"""
        reporte = self.generar_reporte_paquetes_top(fecha_inicio, fecha_fin)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Paquetes Top"
        
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws['A1'] = "RANKING DE PAQUETES MÁS RESERVADOS"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')
        
        ws['A2'] = f"Período: {fecha_inicio} a {fecha_fin}"
        ws['A3'] = f"Fecha de generación: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # Estadísticas
        ws['A5'] = "RESUMEN"
        ws['A5'].font = Font(bold=True, size=11)
        ws['A6'] = f"Paquetes con reservas: {reporte['estadisticas']['total_paquetes']}"
        ws['A7'] = f"Total reservas: {reporte['estadisticas']['total_reservas']}"
        ws['A8'] = f"Total personas: {reporte['estadisticas']['total_personas']}"
        ws['A9'] = f"Ingresos totales: ${reporte['estadisticas']['ingreso_total']:,.2f}"
        
        # Encabezados
        headers = ["Posición", "Paquete", "Destino", "Total Reservas", "Total Personas", "Precio Promedio"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=11, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Datos
        for idx, paquete in enumerate(reporte['datos'], 1):
            row = 11 + idx
            ws.cell(row=row, column=1).value = idx
            ws.cell(row=row, column=2).value = paquete['nombre']
            ws.cell(row=row, column=3).value = paquete['destino_nombre']
            ws.cell(row=row, column=4).value = paquete['total_reservas']
            ws.cell(row=row, column=5).value = paquete['total_personas']
            ws.cell(row=row, column=6).value = paquete['precio_promedio']
            
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                if col == 6:
                    cell.number_format = '$#,##0.00'
        
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        
        excel_io = io.BytesIO()
        wb.save(excel_io)
        excel_io.seek(0)
        
        return excel_io

    def exportar_excel_canceladas(self, fecha_inicio, fecha_fin):
        """Exporta reporte de canceladas a Excel"""
        reporte = self.generar_reporte_canceladas(fecha_inicio, fecha_fin)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Canceladas"
        
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        ws['A1'] = "REPORTE DE RESERVAS CANCELADAS"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:J1')
        
        ws['A2'] = f"Período: {fecha_inicio} a {fecha_fin}"
        ws['A3'] = f"Fecha de generación: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        ws['A5'] = "ESTADÍSTICAS"
        ws['A5'].font = Font(bold=True, size=11)
        ws['A6'] = f"Total canceladas: {reporte['estadisticas']['total_canceladas']}"
        ws['A7'] = f"Valor cancelado: ${reporte['estadisticas']['valor_cancelado']:,.2f}"
        ws['A8'] = f"Promedio por cancelación: ${reporte['estadisticas']['promedio_cancelacion']:,.2f}"
        
        headers = ["ID", "Código", "Fecha Viaje", "Fecha Cancelación", "Cliente", "Teléfono", "Paquete", "Motivo", "Precio"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=11, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for row, cancelada in enumerate(reporte['datos'], 12):
            ws.cell(row=row, column=1).value = cancelada['id_reserva']
            ws.cell(row=row, column=2).value = cancelada['codigo_unico']
            ws.cell(row=row, column=3).value = cancelada['fecha_viaje']
            ws.cell(row=row, column=4).value = cancelada['fecha_reserva']
            ws.cell(row=row, column=5).value = cancelada['cliente_nombre']
            ws.cell(row=row, column=6).value = cancelada['cliente_telefono']
            ws.cell(row=row, column=7).value = cancelada['paquete_nombre']
            ws.cell(row=row, column=8).value = cancelada['motivo']
            ws.cell(row=row, column=9).value = cancelada['precio']
            
            for col in range(1, 10):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                if col == 9:
                    cell.number_format = '$#,##0.00'
        
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 25
        ws.column_dimensions['I'].width = 12
        
        excel_io = io.BytesIO()
        wb.save(excel_io)
        excel_io.seek(0)
        
        return excel_io

    def exportar_excel_ingresos(self, fecha_inicio, fecha_fin):
        """Exporta reporte de ingresos a Excel"""
        reporte = self.generar_reporte_ingresos(fecha_inicio, fecha_fin)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Ingresos"
        
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        ws['A1'] = "REPORTE DE INGRESOS ESPERADOS"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:E1')
        
        ws['A2'] = f"Período: {fecha_inicio} a {fecha_fin}"
        ws['A3'] = f"Fecha de generación: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        ws['A5'] = "RESUMEN GENERAL"
        ws['A5'].font = Font(bold=True, size=11)
        ws['A6'] = f"Total reservas: {reporte['total_reservas']}"
        ws['A7'] = f"Ingresos totales: ${reporte['ingreso_total']:,.2f}"
        ws['A8'] = f"Ingreso promedio por reserva: ${reporte['ingreso_promedio']:,.2f}"
        ws['A9'] = f"Clientes únicos: {reporte['clientes_unicos']}"
        
        # Desglose por estado
        ws['A11'] = "DESGLOSE POR ESTADO"
        ws['A11'].font = Font(bold=True, size=11)
        
        headers = ["Estado", "Reservas", "Ingreso Total", "Ingreso Promedio", "Clientes"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=12, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for row, estado_data in enumerate(reporte['desglose_por_estado'], 13):
            ws.cell(row=row, column=1).value = estado_data['estado']
            ws.cell(row=row, column=2).value = estado_data['reservas']
            ws.cell(row=row, column=3).value = estado_data['ingreso']
            ws.cell(row=row, column=4).value = estado_data['promedio']
            ws.cell(row=row, column=5).value = estado_data['clientes']
            
            for col in range(1, 6):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                if col in [3, 4]:
                    cell.number_format = '$#,##0.00'
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        
        excel_io = io.BytesIO()
        wb.save(excel_io)
        excel_io.seek(0)
        
        return excel_io

    def exportar_excel_destinos_temporada(self, fecha_inicio, fecha_fin):
        """Exporta reporte de destinos por temporada a Excel"""
        reporte = self.generar_reporte_destinos_temporada(fecha_inicio, fecha_fin)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Destinos"
        
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        subheader_fill = PatternFill(start_color="D9E8F5", end_color="D9E8F5", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        ws['A1'] = "REPORTE DE DESTINOS POR TEMPORADA"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:J1')
        
        ws['A2'] = f"Período: {fecha_inicio} a {fecha_fin}"
        ws['A3'] = f"Fecha de generación: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        ws['A5'] = "RESUMEN"
        ws['A5'].font = Font(bold=True, size=11)
        ws['A6'] = f"Total destinos: {reporte['estadisticas']['total_destinos']}"
        ws['A7'] = f"Total reservas: {reporte['estadisticas']['total_reservas']}"
        ws['A8'] = f"Total personas: {reporte['estadisticas']['total_personas']}"
        
        headers = ["Destino", "Mes", "Reservas", "Personas", "Precio Promedio", "Precio Máx", 
                  "Precio Mín", "Paquetes", "Completadas"]
        
        row_actual = 11
        
        for destino in reporte['datos']:
            # Encabezado del destino
            ws.cell(row=row_actual, column=1).value = destino['nombre']
            ws.cell(row=row_actual, column=1).font = Font(bold=True, size=11)
            for col in range(1, 10):
                ws.cell(row=row_actual, column=col).fill = subheader_fill
            row_actual += 1
            
            # Encabezados de columnas
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_actual, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
            row_actual += 1
            
            # Datos de temporadas
            for temporada in destino['temporadas']:
                ws.cell(row=row_actual, column=1).value = ""
                ws.cell(row=row_actual, column=2).value = temporada['mes']
                ws.cell(row=row_actual, column=3).value = temporada['cantidad_reservas']
                ws.cell(row=row_actual, column=4).value = temporada['total_personas']
                ws.cell(row=row_actual, column=5).value = temporada['precio_promedio']
                ws.cell(row=row_actual, column=6).value = temporada['precio_maximo']
                ws.cell(row=row_actual, column=7).value = temporada['precio_minimo']
                ws.cell(row=row_actual, column=8).value = temporada['paquetes_diferentes']
                ws.cell(row=row_actual, column=9).value = temporada['reservas_completadas']
                
                for col in range(1, 10):
                    cell = ws.cell(row=row_actual, column=col)
                    cell.border = border
                    if col in [5, 6, 7]:
                        cell.number_format = '$#,##0.00'
                
                row_actual += 1
            
            row_actual += 1
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 12
        
        excel_io = io.BytesIO()
        wb.save(excel_io)
        excel_io.seek(0)
        
        return excel_io